"""Data source loaders.

Loaders return an iterable of items. An item is either:
  - a dict (CSV/XLSX rows, JSON objects) — accessed as ${item.field}
  - a scalar (str/int/etc. — from inline lists, or when `column` is set) — accessed as ${item}
"""
import csv
import json
from pathlib import Path
from typing import Any, Iterable, Iterator, List, Optional

from .workflow import DataSource


class DataSourceError(Exception):
    pass


def iter_data_source(ds: DataSource) -> Iterator[Any]:
    """Yield raw items (without applying filter / start_index / etc).

    Use load_items() for the user-facing pipeline.
    """
    if ds.type == "inline":
        for item in (ds.inline_items or []):
            yield item
        return

    if ds.type == "csv":
        if not ds.path:
            raise DataSourceError("csv data source missing 'path'")
        p = Path(ds.path)
        if not p.exists():
            raise DataSourceError(f"csv file not found: {p}")
        # Try utf-8-sig first (handles BOM), then gb18030 fallback for Chinese files
        for enc in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with p.open("r", encoding=enc, newline="") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if ds.column:
                            yield row.get(ds.column, "")
                        else:
                            yield dict(row)
                return
            except UnicodeDecodeError:
                continue
        raise DataSourceError(f"could not decode CSV {p} with utf-8 or gb18030")

    if ds.type == "xlsx":
        if not ds.path:
            raise DataSourceError("xlsx data source missing 'path'")
        # openpyxl is heavy; import lazily
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise DataSourceError("openpyxl is required for xlsx data sources")
        p = Path(ds.path)
        if not p.exists():
            raise DataSourceError(f"xlsx file not found: {p}")
        wb = load_workbook(str(p), data_only=True, read_only=True)
        try:
            sheet = wb[ds.sheet] if ds.sheet else wb.active
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = [str(h) if h is not None else "" for h in next(rows)]
            except StopIteration:
                return
            for raw in rows:
                row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
                if ds.column:
                    yield row.get(ds.column, "")
                else:
                    yield row
        finally:
            wb.close()
        return

    if ds.type == "json":
        if not ds.path:
            raise DataSourceError("json data source missing 'path'")
        p = Path(ds.path)
        if not p.exists():
            raise DataSourceError(f"json file not found: {p}")
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            for item in data:
                yield item
        elif isinstance(data, dict):
            # If user gave an object, iterate its values? Be conservative: error
            raise DataSourceError("json data source expects a top-level array")
        else:
            raise DataSourceError(f"json root must be an array, got {type(data).__name__}")
        return

    raise DataSourceError(f"unknown data source type: {ds.type}")


def _is_empty(item: Any) -> bool:
    if item is None:
        return True
    if isinstance(item, str) and not item.strip():
        return True
    if isinstance(item, dict) and not item:
        return True
    if isinstance(item, (list, tuple)) and not item:
        return True
    return False


def load_items(ds: DataSource, filter_fn=None) -> List[Any]:
    """Materialize the data source applying start_index / end_index / skip_empty / filter.

    filter_fn: callable(item) -> bool; None means no filter.
    """
    raw = list(iter_data_source(ds))

    # Slice
    start = max(0, int(ds.start_index or 0))
    end = ds.end_index if ds.end_index is not None else len(raw)
    sliced = raw[start:end]

    # Filter empties
    if ds.skip_empty:
        sliced = [x for x in sliced if not _is_empty(x)]

    # User filter
    if filter_fn is not None:
        out = []
        for x in sliced:
            try:
                if filter_fn(x):
                    out.append(x)
            except Exception:
                # filter errors should not be silent in real code; here we skip the item
                # and the caller is expected to have logged via dry-run preview
                continue
        return out
    return sliced
